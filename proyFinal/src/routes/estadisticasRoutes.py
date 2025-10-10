from flask import (
    Blueprint, request, render_template, redirect, url_for,
    flash, send_file, jsonify, send_from_directory, session
)
import re
import os
import openpyxl
import io 
from datetime import datetime, timezone, date, timedelta
import pytz
import src.controllers.usuarioController as usuarioController
import src.controllers.calendarioController as calendarioController
import src.controllers.estadisticaController as estadisticasController
import src.controllers.notificacionController as notificacionController
import src.utils.enums.generalEnum as generalEnum
from openpyxl.utils import range_boundaries
from src.models.usuario import Usuario
from src.models.estadisticaPorPartido import EstadisticaPorPartido
from src.models.estadisticaUsuarioPartido import EstadisticaUsuarioPartido
from openpyxl.utils import column_index_from_string,range_boundaries
from src import db
from sqlalchemy.engine.row import Row
from src.models.notificacion import Notificacion
from werkzeug.utils import secure_filename
from werkzeug.exceptions import NotFound

estadisticas_bp = Blueprint('estadisticas', __name__, url_prefix='/estadisticas')


@estadisticas_bp.route('/', methods=['GET'])
def index():
    return render_template('estadisticas/index.html')

@estadisticas_bp.route('/ver', methods=['GET'])
def ver():
    categorias = [
        {'value': cat.value, 'text': cat.name}
        for cat in generalEnum.CategoriaEnum
        ]
    ramas = [
        {'value': r.value, 'text': r.name}
        for r in generalEnum.RamaEnum
    ]
    division = [
        {'value': d.value, 'text': d.name}
        for d in generalEnum.DivisionEnum
        ]
    contrincantes = [
        {'value': c.value, 'text': c.name}
        for c in generalEnum.ContrincantesEnum
    ]
    return render_template('estadisticas/ver_estadisticas.html',categorias=categorias, ramas=ramas, division = division, contrincantes=contrincantes)

def rellenar_excel(categoria, rama, division, idPartido, ids_seleccionados=None):
    carpeta_actual = os.path.dirname(__file__)
    ruta_archivo = os.path.abspath(os.path.join(carpeta_actual, "..", "datos", "plantilla_de_estadisticas.xlsx"))
    wb = openpyxl.load_workbook(ruta_archivo)
    ws = wb.active

    partido = calendarioController.getEventoById(int(idPartido))

    if not partido:
        return jsonify({
                "estado": "error",
                "mensaje": f"Partido no encontrado"
            }), 400

    categoria_texto = generalEnum.CategoriaEnum(int(categoria)).name.replace("_", " ").title()
    rama_texto      = generalEnum.RamaEnum(int(rama)).name.replace("_", " ").title()
    division_texto  = generalEnum.DivisionEnum(int(division)).name.replace("_", " ").title()

    # D3
    ws.cell(row=3, column=4, value=f"Rosario Central - {categoria_texto} - {rama_texto} - {division_texto}")

    # O3
    ws.cell(row=3, column=15, value=generalEnum.ContrincantesEnum(int(partido.IdContrincante)).name)  # hardcode por ahora

    # P5
    fecha_str = partido.FechaInicio.strftime("%d-%m-%Y")
    ws.cell(row=5, column=16, value=fecha_str)

    # AE3 -> idPartido 
    ws.cell(row=3, column=31, value=partido.Id)

    usuarios = usuarioController.getUsuarioByCategoriaYRama(categoria, rama, division, ids_seleccionados)
    start_row = 11
    for i, u in enumerate(usuarios, start=start_row):
        ws.cell(row=i, column=1, value=u["Id"])
        ws.cell(row=i, column=2, value=u["Nombre"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="plantilla_estadisticas_rellena.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


columnas_modelo = {
    "REE": "Recepcion E",
    "REV": "Recepcion V",
    "RE0": "Recepcion 0",
    "RE1": "Recepcion 1",
    "RE2": "Recepcion 2",
    "RE3": "Recepcion 3",
    "RETOTAL": "Recepcion Total",

    "ROE": "Rotacion E",
    "ROB": "Rotacion B",
    "RO0": "Rotacion 0",
    "RO1": "Rotacion 1",
    "RO2": "Rotacion 2",
    "RO3": "Rotacion 3",
    "RO4": "Rotacion 4",
    "ROTOTAL": "Rotacion Total",

    "TRE": "Transicion E",
    "TRB": "Transicion B",
    "TR0": "Transicion 0",
    "TR1": "Transicion 1",
    "TR2": "Transicion 2",
    "TR3": "Transicion 3",
    "TR4": "Transicion 4",
    "TRTOTAL": "Transicion Total",

    "SA0": "Saque 0",
    "SA1": "Saque 1",
    "SA2": "Saque 2",
    "SA3": "Saque 3",
    "SA4": "Saque 4",
    "SATOTAL": "Saque Total",

    "BLP": "Bloqueo P",
    "BLN": "Bloqueo N+",
    "BLTOTAL": "Bloqueo Total"
}


def get_cell_value(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            min_col, min_row, _, _ = range_boundaries(str(merged_range))
            return ws.cell(row=min_row, column=min_col).value
    return None


@estadisticas_bp.route('/subir_estadisticas', methods=['POST'])
def subir_estadisticas():
    archivo = request.files.get("archivo")
    fecha_str = request.form.get("fecha")  # dd-mm-YYYY
    idPartido = request.form.get("partido")
    #emails_jugadores_excel = []
    emails_jugadores_excel = {
            "laradelcoro01@gmail.com",
            "lizastrappini99@gmail.com",
            "morakopech@gmail.com",
            "laradelcoro01+3@gmail.com"
        }
    if not archivo or not fecha_str or not idPartido:
        return jsonify({"estado": "error", "mensaje": "Faltan datos"}), 400

    try:
        idPartido = int(idPartido)
    except ValueError:
        return jsonify({"estado": "error", "mensaje": "ID de partido inválido"}), 400

    partido = calendarioController.getEventoById(idPartido)
    if not partido:
        return jsonify({"estado": "error", "mensaje": "No se encontró el partido seleccionado"}), 400

    if partido.TieneEstadistica:
        return jsonify({"estado": "error", "mensaje": "El partido ya tiene estadisticas cargadas"}), 400

    try:
        wb = openpyxl.load_workbook(archivo)
    except Exception:
        return jsonify({"estado": "error", "mensaje": "No se pudo abrir el archivo Excel"}), 400

    ws = wb.active

    # ---------- VALIDACION DE PARTIDO ----------
    id_partido_excel_value = ws['AE3'].value
    if id_partido_excel_value is None:
        return jsonify({"estado": "error", "mensaje": "No se puede determinar el ID del partido en la planilla"}), 400

    try:
        id_partido_excel = int(str(id_partido_excel_value).strip())
    except ValueError:
        return jsonify({"estado": "error", "mensaje": f"El id del partido de la planilla ('{id_partido_excel_value}') no es válido"}), 400

    if idPartido != id_partido_excel:
        return jsonify({"estado": "error", "mensaje": f"El partido seleccionado no coincide con el de la planilla"}), 400

    # ---------- RESULTADO ----------
    resultado_str = ws['AE5'].value
    if resultado_str not in ['G', 'P']:
        return jsonify({"estado": "error", "mensaje": "Resultado del partido inválido, debe ser 'G' o 'P'"}), 400

    try:
        idResultado = generalEnum.ResultadoEnum[resultado_str].value
    except KeyError:
        return jsonify({"estado": "error", "mensaje": "Resultado del partido no reconocido"}), 400

    # ---------- FECHA DESDE EXCEL ----------
    fecha_excel_value = ws['P5'].value

    if not fecha_excel_value:
        return jsonify({
            "estado": "error",
            "mensaje": "No se encontró la fecha en la planilla"
        }), 400

    try:
        fecha = datetime.strptime(fecha_excel_value.strip(), "%d-%m-%Y").date()
    except ValueError:
        return jsonify({
            "estado": "error",
            "mensaje": f"Formato de fecha inválido en la planilla ({fecha_excel_value}), debe ser dd-mm-YYYY"
        }), 400

    # ---------- CREAR ESTADISTICA ----------
    try:
        arg = pytz.timezone("America/Argentina/Buenos_Aires")
        # ---------- CREAR ESTADISTICA ----------
        estadistica = EstadisticaPorPartido(
            Fecha=fecha,
            IdContrincante=partido.IdContrincante,
            IdCategoria=partido.IdCategoria,
            IdRama=partido.IdRama,
            IdDivision=partido.IdDivision,
            Resultado=idResultado,
            IdPartido=partido.Id,
            IdEntrenador=session.get('_user_id'),
            FechaCarga= datetime.now(arg)
        )
        db.session.add(estadistica)

        # ---------- ACTUALIZAR PARTIDO ----------
        partido.TieneEstadistica = True
        db.session.add(partido)  

        # ---------- LEER JUGADORES ----------
        fila = 11
        while True:
            celda_id = ws.cell(row=fila, column=1).value
            if not celda_id:
                break
            try:
                idUsuario = int(celda_id)
            except ValueError:
                raise ValueError(f"ID inválido en fila {fila}")

            usuario = Usuario.query.get(idUsuario)
            if usuario:
                valores = {}
                #emails_jugadores_excel.append(usuario.Email)
                for idx, campo in enumerate(columnas_modelo, start=3):
                    valor = get_cell_value(ws, fila, idx)
                    if valor is None:
                        nombre_legible = columnas_modelo[campo]
                        raise ValueError(f"El campo '{nombre_legible}' está vacío en la fila {fila}")
                    valores[campo] = int(valor)

                rel = EstadisticaUsuarioPartido(
                    IdEstadisticaPorPartido=estadistica.Id,
                    IdUsuario=idUsuario,
                    **valores
                )
                db.session.add(rel)

            fila += 1

        # ---------- COMMIT FINAL ----------
        db.session.commit()

        arg = pytz.timezone("America/Argentina/Buenos_Aires")
        fecha = datetime.now(arg)

        nueva_notif = Notificacion(
            Titulo = "Estadísticas disponibles",
            Descripcion= "Ya podes ver tus estadísticas del partido vs " + generalEnum.ContrincantesEnum(int(partido.IdContrincante)).name,
            IdCategoria= partido.IdCategoria ,
            IdDivision=  partido.IdDivision,
            IdRama= partido.IdRama,
            FechaEnvio=fecha
        )
        notificacionController.agregarNotificacion(nueva_notif)

        for email in emails_jugadores_excel:
            estadisticasController.enviar_mail(
                email,
                "Estadísticas disponibles",
                "Ya podes ver tus estadísticas del partido vs " +
                generalEnum.ContrincantesEnum(int(partido.IdContrincante)).name
            )

        UPLOAD_FOLDER = os.path.join(os.getcwd(), "uploads_estadisticas")
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)

        nombre_archivo = f"estadisticas_{estadistica.Id}.xlsx"
        ruta_guardado = os.path.join(UPLOAD_FOLDER, nombre_archivo)
        archivo.stream.seek(0)     
        archivo.save(ruta_guardado)

        estadistica.RutaArchivo = nombre_archivo
        db.session.add(estadistica)
        db.session.commit()
        return jsonify({"estado": "ok", "mensaje": "Estadísticas cargadas correctamente"})

    except Exception as e:
        db.session.rollback()
        return jsonify({"estado": "error", "mensaje": f"No se pudo guardar los datos: {str(e)}"}), 500
 


@estadisticas_bp.route('/descargar_plantilla', methods=['GET'])
def descargar_plantilla():
    return rellenar_excel()

@estadisticas_bp.route('/archivos', methods=['GET'])
def listar_archivos():
    carpeta = os.path.join('src', 'datos')
    archivos = [f for f in os.listdir(carpeta) if f.lower().endswith(('.xlsx', '.xls'))]
    return jsonify(archivos)


@estadisticas_bp.route('/cargar_estadisticas', methods=['GET'])
def cargar_estadisticas():
    
    categorias = [
    {'value': cat.value, 'text': cat.name}
    for cat in generalEnum.CategoriaEnum
    ]
    ramas = [
        {'value': r.value, 'text': r.name}
        for r in generalEnum.RamaEnum
    ]
    division = [
        {'value': d.value, 'text': d.name}
        for d in generalEnum.DivisionEnum
        ]
    return render_template('estadisticas/index.html',categorias=categorias, ramas=ramas, division = division)


@estadisticas_bp.route('/descargar_planilla', methods=['GET'])
def descargar_planilla():
    categoria = request.args.get("categoria")
    usuarios = request.args.get("usuarios")  
    rama = request.args.get("rama")
    division = request.args.get("division")
    partido = request.args.get("partido")

    if usuarios:
        ids_seleccionados = [int(x) for x in usuarios.split(",") if x.strip()]
    else:
        ids_seleccionados = None

    if not categoria or not rama or not division or not partido:
        flash("No se selecciono categoria", "danger")
        categorias = [
        {'value': cat.value, 'text': cat.name}
        for cat in generalEnum.CategoriaEnum
        ]
        ramas = [
        {'value': r.value, 'text': r.name}
        for r in generalEnum.RamaEnum
        ]

        division = [
        {'value': d.value, 'text': d.name}
        for d in generalEnum.DivisionEnum
        ]
        return render_template('estadisticas/index.html',categorias=categorias, ramas=ramas, division = division)
    
    return rellenar_excel(categoria,rama, division, partido, ids_seleccionados) 


@estadisticas_bp.route('/usuarios_por_categoria', methods=['GET'])
def usuarios_por_categoria():
    categoria = request.args.get("categoria")
    rama = request.args.get("rama")
    division = request.args.get("division")
    
    if not categoria or not rama or not division:
        return jsonify({"estado": "error", "mensaje": "No se seleccionó filtros"}), 400

    usuarios = usuarioController.getUsuarioByCategoriaYRama(categoria, rama, division)

    return jsonify({"estado": "ok", "usuarios": usuarios})

@estadisticas_bp.route('/partidos_por_categoria', methods=['GET'])
def partidos_por_categoria():
    categoria = request.args.get("categoria")
    fecha = request.args.get("fecha")
    rama = request.args.get("rama")
    division = request.args.get("division")

    if not categoria or not fecha:
        return jsonify({"estado": "error", "mensaje": "No se seleccionó categoría"}), 400

    partidos = calendarioController.getPartidosByCategoria(fecha,categoria, rama, division)

    return jsonify({"estado": "ok", "partidos": partidos})

@estadisticas_bp.route('/partidos_por_categoria_mostrar', methods=['GET'])
def partidos_por_categoria_mostrar():
    categoria = request.args.get("categoria")
    fecha = request.args.get("fecha")
    rama = request.args.get("rama")
    division = request.args.get("division")

    if not categoria or not fecha:
        return jsonify({"estado": "error", "mensaje": "No se seleccionó categoría"}), 400

    partidos = calendarioController.getPartidosByCategoriaMostrar(fecha,categoria, rama, division)

    return jsonify({"estado": "ok", "partidos": partidos})


@estadisticas_bp.route('/partidos_por_categoriayfecha', methods=['GET'])
def partidos_por_categoriayfecha():
    categoria = request.args.get("categoria")
    fecha = request.args.get("fecha")

    if not categoria or not fecha:
        return jsonify({"estado": "error", "mensaje": "No se seleccionó categoría"}), 400

    partidos = calendarioController.getPartidosByCategoriaYFecha(fecha,categoria)

    return jsonify({"estado": "ok", "partidos": partidos})


@estadisticas_bp.route('/datos_graficos', methods=['GET'])
def datos_graficos():
    categoria = request.args.get("categoria")
    fecha = request.args.get("fecha")
    rama = request.args.get("rama")
    division = request.args.get("division")
    partido = request.args.get("partido")
    misEstadisticas = request.args.get("misEstadisticas", "false").lower() == "true"
    contrincante = request.args.get("contrincante")
    idUsuario = None

    if(misEstadisticas):
        idUsuario = session.get('_user_id')
        misEstadisticas = True

    if not fecha:
        return jsonify({"estado": "error", "mensaje": "No se seleccionó categoría"}), 400

    partidos = estadisticasController.armarEstadisticas(categoria, rama, division, fecha, partido, idUsuario, contrincante, misEstadisticas)

    return jsonify({"estado": "ok", "partidos": partidos})


@estadisticas_bp.route('/lista', methods=['GET'])
def lista_estadisticas():
    categorias = [
        {'value': cat.value, 'text': cat.name}
        for cat in generalEnum.CategoriaEnum
        ]
    ramas = [
        {'value': r.value, 'text': r.name}
        for r in generalEnum.RamaEnum
    ]
    division = [
        {'value': d.value, 'text': d.name}
        for d in generalEnum.DivisionEnum
        ]
    entrenadores = [
        {'value': u.Id, 'text': f'{u.Nombre} {u.Apellido}'}
        for u in Usuario.query.filter_by(IdRol=3).all()
    ]
    return render_template('estadisticas/listar_estadisticas.html',categorias=categorias, ramas=ramas, division = division, entrenadores=entrenadores)

@estadisticas_bp.route('/listar_estadisticas', methods=['POST'])
def listar_estadisticas():
    categoria = request.form.get("categoria")
    fecha = request.form.get("fecha")
    rama = request.form.get("rama")
    division = request.form.get("division")
    entrenador = request.form.get("entrenador")

    fecha_desde = ''
    fecha_hasta = ''

    if fecha:
        try:
            if " a " in fecha:
                partes = fecha.split(" a ")
                fecha_desde = datetime.strptime(partes[0].strip(), "%d-%m-%Y").date()
                fecha_hasta = datetime.strptime(partes[1].strip(), "%d-%m-%Y").date()
            else:
                fecha_unica = datetime.strptime(fecha.strip(), "%d-%m-%Y").date()
                fecha_desde = fecha_unica
                fecha_hasta = fecha_unica
        except ValueError:
            pass

    data = {
        "categoria": categoria or '',
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "rama": rama or '',
        "division": division or '',
        "entrenador": entrenador or ''
    }

    estadisticas = estadisticasController.obtenerEstadisticasCargadas(data)

    return estadisticas

@estadisticas_bp.route('/eliminar_estadistica', methods=['POST'])
def eliminar_estadistica():
    id = request.form.get("id")
    if not id:
        return jsonify({"estado": "error", "mensaje": "El ID no es válido"}), 400

    resultado = estadisticasController.eliminarEstadistica(id)

    if not resultado:
        return jsonify({"estado": "error", "mensaje": "No se pudo eliminar la estadística"}), 400

    return jsonify({"estado": "ok", "mensaje": "Estadística eliminada con éxito"})


@estadisticas_bp.route('/descargar_estadistica/<int:id>', methods=['GET'])
def descargar_estadistica(id):
    estadistica = EstadisticaPorPartido.query.get(id)
    if not estadistica or not estadistica.RutaArchivo:
        return jsonify({"estado": "error", "mensaje": "Archivo no encontrado"}), 404

    try:
        return send_from_directory(
            "uploads_estadisticas",
            estadistica.RutaArchivo,
            as_attachment=True,
            download_name=f"estadisticas_{id}.xlsx"
        )
    except NotFound:
        return jsonify({"estado": "error", "mensaje": "El archivo no existe en el servidor"}), 404