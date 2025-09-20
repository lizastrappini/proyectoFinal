from datetime import date, datetime
import os
import secrets
import string
from flask import Blueprint, redirect, request, render_template,flash, jsonify, send_file, send_from_directory, url_for
import openpyxl
from src.controllers import deportistaController
from src.controllers.usuarioController import enviar_mail_categoria
from src.models.usuario import Usuario
from werkzeug.security import generate_password_hash
import pytz
from src.utils.enums import generalEnum



deportista_bp = Blueprint('deportista', __name__)


@deportista_bp.route('/')
def index():
    categorias = [
    {'value': cat.value, 'text': cat.name}
    for cat in generalEnum.CategoriaEnum
    ]
    ramas = [
    {'value': rama.value, 'text': rama.name}
    for rama in generalEnum.RamaEnum
    ]
    divisiones = [
    {'value': div.value, 'text': div.name}
    for div in generalEnum.DivisionEnum
    ]
    federados = [
    {'value': federado.value, 'text': federado.name}
    for federado in generalEnum.FederadoEnum
    ]
    localidades = [
    {'value': localidad.value, 'text': localidad.name}
    for localidad in generalEnum.LocalidadEnum
    ]
    deportistas = deportistaController.obtener_deportistas()  
    lista_deportistas = [
        {
            'dni': e['dni'],
            'nombre': e['nombre'],
            'apellido': e['apellido']
        }
        for e in deportistas
    ]
    return render_template('deportista/index.html',localidades=localidades, categorias=categorias,deportistas=lista_deportistas, ramas=ramas, divisiones=divisiones, federados= federados)


@deportista_bp.route('/filtrar')
def filtrar():
    categoria = request.args.get('categoria')
    division = request.args.get('division')
    rama = request.args.get('rama')
    dni = request.args.get('dni')
    
    if categoria and categoria.isdigit():
        categoria = int(categoria)
    
    if division and division.isdigit():
        division = int(division)
    
    if rama and rama.isdigit():
        rama = int(rama)
    
    data = deportistaController.obtener_deportistas(categoria=categoria,dni=dni,rama=rama, division=division) #agregar division=division
    return jsonify({'data': data})




@deportista_bp.route('/nuevoDeportista', methods=['POST'])
def agregar_deportista():
    try:
        dni = request.form.get('dni')
        if not dni or not dni.isdigit() or len(dni) != 8:
            raise ValueError("DNI inválido. Debe contener exactamente 8 dígitos numéricos.")
        nombre = request.form.get('nombre')
        apellido = request.form.get('apellido')
        email = request.form.get('email')
        fechaNacimiento= request.form.get('fechaNacimiento')
        telefono = request.form.get('telefono')
        categoria_nombre = request.form.get('categoria')
        rama_nombre = request.form.get('rama')
        division_nombre = request.form.get('division')
        federado_nombre = request.form.get('federado')
        categoriaExtra = request.form.getlist('categoriaExtra')
        localidad_nombre = request.form.get('localidad')
        categoriaExtraIds = []
        if categoriaExtra:
            try:
                categoriaExtraIds = [
                    generalEnum.CategoriaEnum[nombre].value
                    for nombre in categoriaExtra
                    if nombre in generalEnum.CategoriaEnum.__members__
                ]
            except Exception:
                raise ValueError("Categoría Extra inválida")
            
        # Validar que sean mayores a la categoría principal
        cat_principal_val = generalEnum.CategoriaEnum[categoria_nombre].value
        for cat_extra in categoriaExtraIds:
            if cat_extra <= cat_principal_val:
                raise ValueError("Las categorías extras deben ser mayores a la categoría principal")
            
        # Validar que categoria_nombre esté y sea válido
        if not categoria_nombre or categoria_nombre not in generalEnum.CategoriaEnum.__members__:
            raise ValueError("Categoría inválida o no seleccionada")
        
        if not localidad_nombre or localidad_nombre not in generalEnum.LocalidadEnum.__members__:
            raise ValueError("Localidad inválida o no seleccionada")        
        if not rama_nombre or rama_nombre not in generalEnum.RamaEnum.__members__:
            raise ValueError("Rama inválida o no seleccionada")
        
        if not division_nombre or division_nombre not in generalEnum.DivisionEnum.__members__:
            raise ValueError("División inválida o no seleccionada")
        
        if not fechaNacimiento:
            raise ValueError("La fecha de nacimiento es requerida")

        fecha_nacimiento_dt = datetime.strptime(fechaNacimiento, "%Y-%m-%d")  # Convertir string a datetime
        #categoria_id = deportistaController.calcular_categoria_por_fecha(fecha_nacimiento_dt)
        #categoria_id = generalEnum.CategoriaEnum[categoria_nombre].value
        categoria_id_seleccionada = generalEnum.CategoriaEnum[categoria_nombre].value
        categoria_id_calculada = deportistaController.calcular_categoria_por_fecha(fecha_nacimiento_dt)

        

        if categoria_id_seleccionada != categoria_id_calculada:
            raise ValueError(f"La categoría seleccionada ({categoria_nombre}) no corresponde con la edad del deportista. Debería ser {generalEnum.CategoriaEnum(categoria_id_calculada).name}")

        rama_id = generalEnum.RamaEnum[rama_nombre].value
        division_id = generalEnum.DivisionEnum[division_nombre].value
        federado_id = generalEnum.FederadoEnum[federado_nombre].value
        localidad_id = generalEnum.LocalidadEnum[localidad_nombre].value
        # caracteres = string.ascii_letters + string.digits  # letras + números
        # password_plana = ''.join(secrets.choice(caracteres) for _ in range(8))  

        usuario_existente = Usuario.query.filter_by(Dni=dni).first()
        if usuario_existente:
         raise ValueError(f"Ya existe un usuario con el DNI {dni}")
     
        mail_usuario = Usuario.query.filter_by(Email=email).first()
        if mail_usuario:
         raise ValueError(f"Ya existe un usuario con el mismo email")
        
        arg = pytz.timezone("America/Argentina/Buenos_Aires")

        nuevo_deportista = Usuario(
            Dni= dni,
            Nombre=nombre,
            Apellido=apellido,
            Email= email,
            FechaNacimiento= fecha_nacimiento_dt,
            IdCategoria = categoria_id_seleccionada,
            IdRama = rama_id,
            IdDivision = division_id,
            Password = generate_password_hash(dni),
            NombreUsuario=f"{nombre}_{dni}",
            Localidad= localidad_id,
            IdEstado=1,
            #Direccion="N/A",
            Telefono=telefono,
            IdRol=2,
            Token=None,
            TokenEnviado=False,
            FechaVencimientoToken=None,
            Federado = federado_id,
            CategoriaExtra = ",".join(map(str, categoriaExtraIds)) if categoriaExtraIds else None,
            FechaAlta = datetime.now(arg)
        )
        deportistaController.agregarDeportista(nuevo_deportista)
        deportistaController.enviar_mail_alta_deportista(nuevo_deportista, dni)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': 'Deportista creado exitosamente'}), 200
        else:
            flash('Deportista creado exitosamente', 'success')
            return redirect(url_for('deportista.index'))
    except Exception as e:
        mensaje_error = str(e)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': mensaje_error}), 400
        else:
            flash(f'Error al crear deportista: {mensaje_error}', 'danger')
            return redirect(url_for('deportista.index'))




@deportista_bp.route('/editar/<int:dni>', methods=['POST'])
def editar_deportista(dni):
    try:
        nuevo_dni = request.form.get('dni')
        nuevo_email = request.form.get('email')
        categoria = request.form.get('categoria')
        categoria_extra = request.form.getlist('categoriaExtra')
        nombre = request.form.get('nombre')
        apellido = request.form.get('apellido')
        fechaNacimiento = request.form.get('fechaNacimiento')
        telefono = request.form.get('telefono')
        
        fecha_nacimiento_dt = datetime.strptime(fechaNacimiento, "%Y-%m-%d")
        categoria_id_calculada = deportistaController.calcular_categoria_por_fecha(fecha_nacimiento_dt)
        categoria_id_seleccionada = generalEnum.CategoriaEnum[categoria].value
        # nueva_cat = request.form.get('categoria')
        rama_nombre = request.form.get('rama')
        division_nombre = request.form.get('division')
        federado_nombre = request.form.get('federado')
        localidad_nombre = request.form.get('localidad')
        # categoria_extra = request.form.get('categoriaExtra')
        # categoria = request.form.get('categoria')

        categoriaExtraIds = []
        if categoria_extra:
            try:
                categoriaExtraIds = [
                    generalEnum.CategoriaEnum[nombre].value
                    for nombre in categoria_extra
                    if nombre in generalEnum.CategoriaEnum.__members__
                ]
            except Exception:
                raise ValueError("Categoría Extra inválida")
        
            # Validar que sean mayores a la categoría principal
            cat_principal_val = generalEnum.CategoriaEnum[categoria].value
            for cat_extra in categoriaExtraIds:
                if cat_extra <= cat_principal_val:
                    raise ValueError("Las categorías extra deben ser mayores a la categoría principal")

        if nuevo_dni and int(nuevo_dni) != dni:
            usuario_existente = Usuario.query.filter_by(Dni=nuevo_dni).first()
            if usuario_existente:
                raise ValueError(f"Ya existe un usuario con el DNI {nuevo_dni}")
        if not nuevo_dni or not nuevo_dni.isdigit() or len(nuevo_dni) != 8:
            raise ValueError("DNI inválido. Debe contener exactamente 8 dígitos numéricos.")
        
        deportista = deportistaController.obtener_deportista_por_dni(dni)
        
        if not deportista:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'Deportista no encontrado'}), 400
            else:
                flash('Deportista no encontrado', 'danger')
                return redirect(url_for('deportista.index'))
        
        if nuevo_email and nuevo_email != deportista.Email:
            email_usuario = Usuario.query.filter_by(Email=nuevo_email).first()
            if email_usuario:
                raise ValueError(f"Ya existe un usuario con el mismo email")
        

        if categoria_id_seleccionada != categoria_id_calculada:
            raise ValueError(
                f"La categoría seleccionada ({categoria}) no corresponde con la edad del deportista. "
                f"Debería ser {generalEnum.CategoriaEnum(categoria_id_calculada).name}"
                )
    
        # Actualiza campos
        deportista.Dni = nuevo_dni
        deportista.Nombre = nombre
        deportista.Apellido = apellido
        deportista.Email = nuevo_email
        deportista.FechaNacimiento = fechaNacimiento
        deportista.Telefono = telefono
        deportista.IdCategoria = generalEnum.CategoriaEnum[categoria].value
        # fecha_nacimiento_dt = datetime.strptime(fechaNacimiento, "%Y-%m-%d")
        #deportista.IdCategoria = deportistaController.calcular_categoria_por_fecha(fecha_nacimiento_dt)
        deportista.IdRama = generalEnum.RamaEnum[rama_nombre].value
        deportista.IdDivision = generalEnum.DivisionEnum[division_nombre].value
        deportista.Federado = generalEnum.FederadoEnum[federado_nombre].value
        # deportista.CategoriaExtra = generalEnum.CategoriaEnum[categoria_extra].value
        deportista.CategoriaExtra = ",".join(map(str, categoriaExtraIds)) if categoriaExtraIds else None
        deportista.Localidad = generalEnum.LocalidadEnum[localidad_nombre].value
         
        categoria_vieja = deportista.IdCategoria
        
        if deportista.IdCategoria != categoria_vieja:
            enviar_mail_categoria(
                deportista.Email,
                deportista.Nombre,
                generalEnum.CategoriaEnum(deportista.IdCategoria).name
            )

        deportistaController.actualizar_deportista(deportista)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': 'Deportista actualizado exitosamente'}), 200
        else:
            flash('Deportista actualizado exitosamente', 'success')
            return redirect(url_for('deportista.index'))

    except ValueError as ve:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': str(ve)}), 400
        else:
            flash(str(ve), 'danger')
            return redirect(url_for('deportista.index'))

    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'Error inesperado: {str(e)}'}), 500
        else:
            flash('Ocurrió un error inesperado', 'danger')
            return redirect(url_for('deportista.index'))





@deportista_bp.route('/eliminar/<int:dni>', methods=['POST'])
def eliminar_deportista(dni):
    deportista = deportistaController.obtener_deportista_por_dni(dni)
    if not deportista:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Deportista no encontrado'}), 200
        else:
            flash('Deportista no encontrado', 'danger')
            return redirect(url_for('deportista.index'))
    
    deportistaController.borrar_deportista(deportista)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'message': 'Entrenador eliminado'})
    else:
        flash('Deportista eliminado exitosamente', 'success')
        return redirect(url_for('deportista.index'))
    
    
@deportista_bp.route('/cambiarEstado/<int:dni>', methods=['POST'])
def cambiar_estado(dni):
    deportista = deportistaController.obtener_deportista_por_dni(dni)

    if not deportista:
        return jsonify({'success': False, 'message': 'Deportista no encontrado'}), 404

    # Alternar estado
    if int(deportista.IdEstado) == generalEnum.EstadoEnum.Activo:
        deportista.IdEstado = generalEnum.EstadoEnum.Inactivo
    else:
        deportista.IdEstado = generalEnum.EstadoEnum.Activo

    deportistaController.actualizar_deportista(deportista)

    return jsonify({
        'success': True,
        'message': 'Estado actualizado',
        'nuevo_estado': generalEnum.EstadoEnum(int(deportista.IdEstado)).name,
   
    })


@deportista_bp.route('/getDeportista/<int:dni>', methods=['GET'])
def getDeportista(dni):
    deportista = deportistaController.obtener_deportista_por_dni(dni)

    if not deportista:
        return jsonify({'error': 'Deportista no encontrado'}), 404

    return jsonify({
        "dni": deportista.Dni,
        "nombre": deportista.Nombre,
        "apellido": deportista.Apellido,
        "email": deportista.Email,
        "telefono": deportista.Telefono,
        # devolvemos los nombres porque tus selects usan .text
        "categoria": generalEnum.CategoriaEnum(deportista.IdCategoria).name if deportista.IdCategoria else None,
        "localidad": generalEnum.LocalidadEnum(int(deportista.Localidad)).name if deportista.Localidad else None,
        "division": generalEnum.DivisionEnum(deportista.IdDivision).name if deportista.IdDivision else None,
        "rama": generalEnum.RamaEnum(deportista.IdRama).name if deportista.IdRama else None,
        "fechaNacimiento": deportista.FechaNacimiento.strftime("%Y-%m-%d") if deportista.FechaNacimiento else None,
        "federado": generalEnum.FederadoEnum(deportista.Federado).name if deportista.Federado else None,
        # categorías extra como lista de nombres
        "categoriaExtra": [
            generalEnum.CategoriaEnum(int(x)).name for x in deportista.CategoriaExtra.split(",")
        ] if deportista.CategoriaExtra else []
    })
    

    
@deportista_bp.route('/descargar_planilla')
def descargar_planilla():
    carpeta = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "datos"))
    return send_from_directory(carpeta, "plantilla_deportista.xlsx", as_attachment=True)

       
        
@deportista_bp.route('/importar_deportistas', methods=['POST'])
def importar_deportistas():
    try:
        archivo = request.files.get("archivoExcel")
        if not archivo:
            raise ValueError("Debe subir un archivo Excel")
            
        
        # ---------- ABRIR ARCHIVO ----------
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
        except Exception:
            raise ValueError("No se pudo abrir el archivo Excel")

        # ---------- LEER FILAS ----------
        fila = 2
        registros_creados = 0
        errores = []
        if not ws.cell(row=2, column=1).value:
            raise ValueError("El archivo se encuentra vacío")

        while True:
            dni = ws.cell(row=fila, column=1).value
            
            if not dni:  # fin del archivo
                break
                    
          
            try:
                nombre_val = ws.cell(row=fila, column=2).value
                apellido_val = ws.cell(row=fila, column=3).value
                email_val = ws.cell(row=fila, column=4).value
                fecha_nac_val = ws.cell(row=fila, column=5).value
                telefono_val = ws.cell(row=fila, column=6).value
                categoria_val = ws.cell(row=fila, column=7).value
                rama_val = ws.cell(row=fila, column=8).value
                division_val = ws.cell(row=fila, column=9).value
                federado_val = ws.cell(row=fila, column=10).value
                categoria_extra_val = ws.cell(row=fila, column=11).value
                localidad_val = ws.cell(row=fila, column=12).value

                # ---------- Validaciones ----------
                # DNI
                # if not dni or not str(dni).isdigit() or len(str(dni)) != 8:
                #     raise ValueError(f"DNI inválido en fila {fila}")

                usuario_existente = Usuario.query.filter_by(Dni=dni).first()
                if usuario_existente:
                    raise ValueError(f"Ya existe un usuario con el DNI {dni}")

                # Email duplicado
                mail_usuario = Usuario.query.filter_by(Email=email_val).first()
                if mail_usuario:
                    raise ValueError(f"Ya existe un usuario con el mismo email")

                # Fecha
                if not fecha_nac_val:
                    raise ValueError("La fecha de nacimiento es obligatoria")

                if isinstance(fecha_nac_val, str):
                    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
                        try:
                            fecha_nac = datetime.strptime(fecha_nac_val, fmt)
                            break
                        except ValueError:
                            fecha_nac = None
                    if not fecha_nac:
                        raise ValueError(f"Formato de fecha inválido en fila {fila}. Use DD-MM-YYYY o YYYY-MM-DD")
                elif isinstance(fecha_nac_val, datetime):
                    fecha_nac = fecha_nac_val
                elif isinstance(fecha_nac_val, date):
                    fecha_nac = datetime.combine(fecha_nac_val, datetime.min.time())
                else:
                    raise ValueError(f"Fecha de nacimiento inválida en fila {fila}")

                # Enums obligatorios
                if not categoria_val or categoria_val not in generalEnum.CategoriaEnum.__members__:
                    raise ValueError(f"Categoría inválida en fila {fila}")
                categoria_id = generalEnum.CategoriaEnum[categoria_val].value

                if not rama_val or rama_val not in generalEnum.RamaEnum.__members__:
                    raise ValueError(f"Rama inválida en fila {fila}")
                rama_id = generalEnum.RamaEnum[rama_val].value

                if not division_val or division_val not in generalEnum.DivisionEnum.__members__:
                    raise ValueError(f"División inválida en fila {fila}")
                division_id = generalEnum.DivisionEnum[division_val].value

                if not federado_val or federado_val not in generalEnum.FederadoEnum.__members__:
                    raise ValueError(f"Federado inválido en fila {fila}")
                federado_id = generalEnum.FederadoEnum[federado_val].value

                if not localidad_val or localidad_val not in generalEnum.LocalidadEnum.__members__:
                    raise ValueError(f"Localidad inválida en fila {fila}")
                localidad_id = generalEnum.LocalidadEnum[localidad_val].value

                # Categoria Extra (puede venir como string separado por comas)
                categoriaExtraIds = []
                if categoria_extra_val:
                    try:
                        nombres_extra = [n.strip() for n in str(categoria_extra_val).split(",")]
                        categoriaExtraIds = [
                            generalEnum.CategoriaEnum[nombre].value
                            for nombre in nombres_extra
                            if nombre in generalEnum.CategoriaEnum.__members__
                        ]
                    except Exception:
                        raise ValueError(f"Categoría Extra inválida en fila {fila}")
                #validar que las cat extras no sean menores a la cat principal
                cat_principal_val = generalEnum.CategoriaEnum[categoria_val].value
                for cat_extra in categoriaExtraIds:
                    if cat_extra <= cat_principal_val:
                        raise ValueError("Las categorías extras deben ser mayores a la categoría principal")
                
                categoria_id_seleccionada = generalEnum.CategoriaEnum[categoria_val].value
                categoria_id_calculada = deportistaController.calcular_categoria_por_fecha(fecha_nac_val)

                

                if categoria_id_seleccionada != categoria_id_calculada:
                    raise ValueError(f"La categoría seleccionada ({categoria_val}) no corresponde con la edad del deportista. Debería ser {generalEnum.CategoriaEnum(categoria_id_calculada).name}")
                
                arg = pytz.timezone("America/Argentina/Buenos_Aires")

                # ---------- Crear objeto ----------
                nuevo_deportista = Usuario(
                    Dni=dni,
                    FechaNacimiento=fecha_nac,
                    Nombre=nombre_val,
                    Apellido=apellido_val,
                    Email=email_val,
                    Telefono=telefono_val,
                    Localidad=localidad_id,
                    IdCategoria=categoria_id_seleccionada,
                    IdRama=rama_id,
                    IdDivision=division_id,
                    CategoriaExtra=",".join(map(str, categoriaExtraIds)) if categoriaExtraIds else None,
                    Federado=federado_id,
                    Password=generate_password_hash(dni),
                    NombreUsuario=f"{nombre_val}_{dni}",
                    IdEstado=1,
                    IdRol=2,
                    Token=None,
                    TokenEnviado=False,
                    FechaVencimientoToken=None,
                    FechaAlta = datetime.now(arg)
                    
                )

                deportistaController.agregarDeportista(nuevo_deportista)
                deportistaController.enviar_mail_alta_deportista(nuevo_deportista, dni)
                registros_creados += 1

            except Exception as e:
                errores.append(f"Fila {fila}: {str(e)}")

            fila += 1

            # ---------- RESPUESTA ----------
            if errores:
                preview = errores[:5]
                extra = len(errores) - len(preview)
                mensaje = f"Se importaron {registros_creados} deportistas.\nErrores:\n" + "\n".join(preview)
                if extra > 0:
                    mensaje += f"\n... y {extra} más"

                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'message': mensaje}), 400
                else:
                    flash(mensaje, 'danger')
                    return redirect(url_for('deportista.index'))

            mensaje_ok = f"Se importaron {registros_creados} deportistas correctamente"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': True, 'message': mensaje_ok}), 200
            else:
                flash(mensaje_ok, 'success')
                return redirect(url_for('deportista.index'))

    except Exception as e:
        mensaje_error = str(e)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': mensaje_error}), 400
        else:
            flash(f"Error al importar deportistas: {mensaje_error}", 'danger')
            return redirect(url_for('deportista.index'))
