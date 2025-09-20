from datetime import datetime, timezone, date, timedelta
import pytz
from decimal import Decimal
import secrets
from flask import Blueprint, redirect, request, render_template,flash, jsonify, send_from_directory, url_for
import openpyxl
from src.controllers import deportistaController, pagosController, parametroController
from src.models.parametro import Parametro
from src.models.usuario import Usuario
from src.models.pago import Pago
from werkzeug.security import generate_password_hash
import os
from werkzeug.utils import secure_filename
from collections import Counter
import locale, calendar

from src.utils.enums import generalEnum


pago_bp = Blueprint('pago', __name__)
# deportista_bp = Blueprint('deportista', __name__)


@pago_bp.route('/')
def index():
    estados = [
    {'value': estado.value, 'text': estado.name}
    for estado in generalEnum.EstadoPagoEnum
    ]
    deportistas = Usuario.query.filter_by(IdRol=2).all()
    lista_deportistas = [
        {'id': d.Id, 'nombre': f'{d.Nombre} {d.Apellido}'}
        for d in deportistas
    ]
    return render_template('pago/index.html', estados=estados, deportistas=lista_deportistas)  

@pago_bp.route('/obtener', methods=['GET'])
def obtener():
    pagos = pagosController.obtener_pagos()
    return jsonify(data=pagos)

  

@pago_bp.route('/filtrar')
def filtrar():
    estado = request.args.get('estado')
    fecha_desde = request.args.get('fechaDesde')
    fecha_hasta = request.args.get('fechaHasta')
    filtrado_manual = request.args.get('filtrado_manual', 'false') == 'true'  # nuevo
    # ⚠️ Solo mostrar mensaje si el filtrado es manual
    if not fecha_desde or not fecha_hasta:
        if filtrado_manual:
            return jsonify({'data': [], 'message': 'Debe seleccionar ambas fechas'})
        else:
            return jsonify({'data': [], 'message': '', 'estadisticas': {}})  # tabla vacía sin alerta

    if estado and estado.isdigit():
        estado = int(estado)

    # --- 1) Data para la tabla ---
    data = pagosController.obtener_pagos(
        estado=estado,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta
    )

    # --- 2) Data para estadísticas ---
    fecha_desde_dt = datetime.strptime(fecha_desde, "%Y-%m-%d")
    fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)

    pagos_filtrados = Pago.query.filter(
        Pago.FechaPago >= fecha_desde_dt.date(),
        Pago.FechaPago <= fecha_hasta_dt.date(),
        Pago.IdEstado == generalEnum.EstadoPagoEnum.Pago.value  # ✅ solo pagados
    ).all()

    # ✅ Cantidad de cuotas pagas por categoría (mostrar todas aunque tengan 0)
    cuotas_por_categoria = []
    for cat in generalEnum.CategoriaEnum:
        if cat.value == 0:  # 🚫 saltar NoEspecificada
            continue
        cantidad = sum(
            1 for p in pagos_filtrados
            if p.usuario and p.usuario.IdCategoria == cat.value
        )
        cuotas_por_categoria.append({"nombre": cat.name, "cantidad": cantidad})

    # ✅ Cuotas pagas por mes (año actual completo)
    ahora = datetime.now()
    year = ahora.year
    cuotas_por_mes = [0] * 12

    pagos_ano = Pago.query.filter(
        Pago.FechaPago >= date(year, 1, 1),
        Pago.FechaPago <= date(year, 12, 31),
        Pago.IdEstado == generalEnum.EstadoPagoEnum.Pago.value
    ).all()

    for p in pagos_ano:
        if p.FechaPago:
            cuotas_por_mes[p.FechaPago.month - 1] += 1

    # ✅ Comparativa mes actual vs anterior
    max_mes_index = max(range(12), key=lambda i: cuotas_por_mes[i])
    max_mes_total = cuotas_por_mes[max_mes_index]
    

    # mes anterior al de máximo
    mes_anterior_index = max_mes_index - 1 if max_mes_index > 0 else 11
    total_mes_anterior = cuotas_por_mes[mes_anterior_index]

    if total_mes_anterior > 0:
        variacion = ((max_mes_total - total_mes_anterior) / total_mes_anterior) * 100
    elif max_mes_total > 0:
        variacion = 100.0  # antes era 0, ahora hay pagos
    else:
        variacion = 0.0
    
    MESES_ES = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }

    comparativa = {
        "mes_actual": MESES_ES[max_mes_index + 1],
        "total_actual": max_mes_total,
        "mes_anterior": MESES_ES[mes_anterior_index + 1],
        "total_anterior": total_mes_anterior,
        "variacion_pct": variacion
    }

    return jsonify({
        "data": data,
        "estadisticas": {
            "totalPagas": sum(c["cantidad"] for c in cuotas_por_categoria),
            "porCategoria": cuotas_por_categoria,   # 🔹 siempre todas las categorías
            "porMes": cuotas_por_mes,               # 🔹 12 números
            "mesMax": {
                "nombre": comparativa["mes_actual"],
                "comparacion": f'{comparativa["variacion_pct"]:.1f}%',
                "positivo": comparativa["variacion_pct"] >= 0,
                "comparadoCon": comparativa["mes_anterior"]
            }
        }
    })

@pago_bp.route('/nuevoPago', methods=['POST'])
def agregar_pago():
    try:
        
        fechaPago_str = request.form.get('fechaPago')
        # fechaVencimiento = request.form.get('fechaVencimiento')
        # importe = request.form.get('importe')
        estado_nombre = request.form.get('estado')
        usuario_id = request.form.get('deportista')
        
        if fechaPago_str:
            fechaPago = datetime.strptime(fechaPago_str, "%Y-%m-%d")
        else:
            fechaPago = None 
        if not usuario_id:
            raise ValueError("Debe seleccionar un deportista")
        
        if not estado_nombre or estado_nombre not in generalEnum.EstadoPagoEnum.__members__:
            raise ValueError("Estado inválido o no seleccionada")
        
        parametro = Parametro.query.filter_by(Titulo='ValorCuota').first()
        if not parametro:
            raise ValueError("No se encontró el parámetro ValorCuota")

        try:
            importe = float(parametro.Valor)
        except ValueError:
            raise ValueError("El valor de ValorCuota no es un número válido")
        
        fechaVencimiento_str = request.form.get('fechaVencimiento')
        if fechaVencimiento_str:
            fechaVencimiento = datetime.strptime(fechaVencimiento_str, "%Y-%m-%d")
        else:
            fechaVencimiento = None

  


        estado_id = generalEnum.EstadoPagoEnum[estado_nombre].value
        nuevo_pago = Pago(
            FechaPago= fechaPago,
            FechaVencimiento=fechaVencimiento,
            Importe=importe,
            IdEstado= estado_id,
            IdUsuario =int(usuario_id)
        )
        pagosController.agregarPago(nuevo_pago)
        

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': 'Pago creado exitosamente'}), 200
        else:
            flash('Pago creado exitosamente', 'success')
            return redirect(url_for('pago.index'))
    except Exception as e:
        mensaje_error = str(e)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': mensaje_error}), 400
        else:
            flash(f'Error al crear pago: {mensaje_error}', 'danger')
            return redirect(url_for('pago.index'))
  
    
@pago_bp.route('/importar_pagos', methods=['POST'])
def importar_pagos():
    try:
        archivo = request.files.get("archivoExcel")
        if not archivo:
            raise ValueError("Debe subir un archivo Excel")
        
        # ---------- ABRIR ARCHIVO ----------
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
        except Exception:
            raise ValueError("No se pudo abrir el archivo Excel")

        # ---------- VALIDAR ENCABEZADOS ----------
        encabezados = ["DNI Deportista", "Fecha Pago", "Fecha Vencimiento", "IdEstado"]
        for idx, esperado in enumerate(encabezados, start=1):
            valor = str(ws.cell(row=1, column=idx).value).strip() if ws.cell(row=1, column=idx).value else ""
            if valor != esperado:
                raise ValueError(f"El encabezado en la columna {idx} debe ser '{esperado}'")

        # ---------- LEER FILAS ----------
        fila = 2
        registros_creados = 0
        errores = []

        while True:
            dni = ws.cell(row=fila, column=1).value
            if not dni:  # fin del archivo
                break
            
            usuario = Usuario.query.filter_by(Dni=str(dni).strip()).first()
            if not usuario:
                raise ValueError(f"Fila {fila}: No existe un usuario con DNI {dni}")
                fila += 1
                continue
            parametro = Parametro.query.filter_by(Titulo='ValorCuota').first()
            if not parametro:
                raise ValueError("No se encontró el parámetro ValorCuota")

            try:
                importe = float(parametro.Valor)
            except ValueError:
                raise ValueError("El valor de ValorCuota no es un número válido")
            try:
                fecha_pago_val = ws.cell(row=fila, column=2).value
                fecha_venc_val = ws.cell(row=fila, column=3).value
                # importe_val = ws.cell(row=fila, column=4).value
                id_estado = ws.cell(row=fila, column=4).value

              
                if not id_estado:
                    raise ValueError("Estado inválido o no seleccionado")
                if not fecha_venc_val:
                    raise ValueError("Las fechas son obligatorias")
                # if not importe_val:
                #     raise ValueError("El importe es obligatorio")

                # Parseo fechas si vienen en string
                fecha_pago = fecha_pago_val
                fecha_vencimiento = fecha_venc_val
                if isinstance(fecha_pago_val, str):
                    fecha_pago = datetime.datetime.strptime(fecha_pago_val, "%d-%m-%Y")
                if isinstance(fecha_venc_val, str):
                    fecha_vencimiento = datetime.datetime.strptime(fecha_venc_val, "%d-%m-%Y")

                # Importe
                # importe = Decimal(str(importe_val))

                nuevo_pago = Pago(
                    FechaPago=fecha_pago,
                    FechaVencimiento=fecha_vencimiento,
                    Importe=importe,
                    IdEstado=int(id_estado),
                    IdUsuario=usuario.Id
                )
                pagosController.agregarPago(nuevo_pago)
                registros_creados += 1

            except Exception as e:
                errores.append(f"Fila {fila}: {str(e)}")

            fila += 1

        # ---------- RESPUESTA ----------
        if errores:
            mensaje = f"Se importaron {registros_creados} pagos, con errores en algunas filas: {errores}"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': mensaje}), 400
            else:
                flash(mensaje, 'danger')
                return redirect(url_for('pago.index'))

        mensaje_ok = f"Se importaron {registros_creados} pagos correctamente"
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': mensaje_ok}), 200
        else:
            flash(mensaje_ok, 'success')
            return redirect(url_for('pago.index'))

    except Exception as e:
        mensaje_error = str(e)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': mensaje_error}), 400
        else:
            flash(f"Error al importar pagos: {mensaje_error}", 'danger')
            return redirect(url_for('pago.index'))
        
@pago_bp.route('/editar/<int:id>', methods=['POST'])
def editar_pago(id):
    pago = pagosController.obtener_pago_por_id(id)
    if not pago:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Pago no encontrado'}), 400
        else:
            flash('Pago no encontrado', 'danger')
            return redirect(url_for('pago.index'))

    fechaPago_str = request.form.get('fechaPago')
    fechaVencimiento = request.form.get('fechaVencimiento')
    # importe = request.form.get('importe')
    estado_nombre = request.form.get('estado')
    usuario_id = request.form.get('deportista')
    
    if fechaPago_str:
            fechaPago = datetime.strptime(fechaPago_str, "%Y-%m-%d")
    else:
            fechaPago = None 
    # Actualiza campos
    pago.FechaPago = fechaPago
    pago.FechaVencimiento = fechaVencimiento
    # pago.Importe = importe
    pago.IdEstado = generalEnum.EstadoPagoEnum[estado_nombre].value
    pago.IdUsuario= usuario_id
    

    pagosController.actualizar_pago(pago)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'message': 'Pago actualizado exitosamente'}), 200
    else:
        flash('Pago actualizado exitosamente', 'success')
        return redirect(url_for('pago.index'))



@pago_bp.route('/eliminar/<int:id>', methods=['POST'])
def eliminar_pago(id):
    pago = pagosController.obtener_pago_por_id(id)
    if not pago:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Pago no encontrado'}), 200
        else:
            flash('Pago no encontrado', 'danger')
            return redirect(url_for('pago.index'))
    
    pagosController.borrar_pago(pago)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'message': 'Pago eliminado'})
    else:
        flash('Pago eliminado exitosamente', 'success')
        return redirect(url_for('pago.index'))
    
    

@pago_bp.route('/pagar_seleccionados', methods=['POST'])
def pagar_seleccionados():
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'success': False, 'message': 'No se seleccionaron pagos'}), 400
        
        pagos = Pago.query.filter(Pago.Id.in_(ids)).all()

        actualizados = 0

        for pago in pagos:
            if pago.IdEstado == 1: #ya estaba pagado,para q no actualice fechapago
                continue
            pago.IdEstado = 1
            arg = pytz.timezone("America/Argentina/Buenos_Aires")
            pago.FechaPago = datetime.now(arg)
            pagosController.actualizar_pago(pago)
            actualizados += 1

        mensaje = f'{actualizados} pagos actualizados a Pago'
    
        return jsonify({'success': True,'updated': actualizados,'message': mensaje})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@pago_bp.route('/actualizar_cuota', methods=['POST'])
def actualizar_cuota():
    try:
        # Obtener el nuevo valor desde el formulario o JSON
        nuevo_valor = request.form.get('importe')  # o request.json.get('nuevo_valor')
        if not nuevo_valor:
            raise ValueError("Debe indicar el nuevo valor de la cuota")

        # Buscar el parámetro ValorCuota
        parametro = Parametro.query.filter_by(Titulo='ValorCuota').first()
        if not parametro:
            raise ValueError("No se encontró el parámetro ValorCuota")

        # Actualizar el valor
        parametro.Valor = str(nuevo_valor)  # ⚠️ convertir a string si es Text
       
        parametroController.actualizar_parametro(parametro)
       
        deportistas = Usuario.query.filter_by(IdRol=2).all()  # rol 2 = deportista
        enviados = parametroController.enviar_mail_actualizacion(deportistas, parametro)

        mensaje = f'Cuota actualizada exitosamente. Correos enviados a {enviados} deportistas.'
        

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': mensaje}), 200
        else:
            flash('Cuota actualizada exitosamente', 'success')
            return redirect(url_for('pago.index'))
    except Exception as e:
        mensaje_error = str(e)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': mensaje_error}), 400
        else:
            flash(f'Error al actualizar cuota: {mensaje_error}', 'danger')
            return redirect(url_for('pago.index'))
        

@pago_bp.route('/descargar_planilla')
def descargar_planilla():
    carpeta = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "datos"))
    return send_from_directory(carpeta, "planilla_pagos.xlsx", as_attachment=True)